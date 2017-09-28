import cv2
import os
import numpy as np
import copy
from openpyxl import Workbook
from openpyxl import load_workbook
import xlrd
import xlwt
import json

image_rows = 424
image_cols = 512
data_path = 'images/trial_2/p3+5/middle/'
image_prefix = 'p3s1d_'
image_path = 'images/'
label_path = 'annotations/'


# depth_image = cv2.imread(os.path.join(data_path, '2000.png'),cv2.IMREAD_UNCHANGED)
# depth_image = depth_image.astype('float32')
# # depth_image *= (255.0/2500.0)
# # print(depth_image[200])   
# # depth_image = depth_image.astype('int8')
# # depth_image = cv2.convertScaleAbs(depth_image, alpha=(1/2700))
# print(depth_image[10]/2500.0)
#   
# # M_rotate  = cv2.getRotationMatrix2D((512/2,424/2),180,1)
# # result = cv2.warpAffine(depth_image, M_rotate, (512, 424))
# result = depth_image.T
# flipped = cv2.flip(result,0)
# result = flipped[5:405,20:220]
# cv2.imshow('Labeling', result)
# cv2.waitKey(0)    

# a = [1,0,0]
# b = [9,10,0]
# x1 = np.where(a == np.max(a))
# x2 = np.where(b == np.max(b))
# print(b.index(max(b)))


# 
# import xlrd
# from pyExcelerator import *
#   
# w = Workbook() 
# ws = w.add_sheet('Sheet1') 
#  
# fname = "result.xlsx"
# bk = xlrd.open_workbook(fname)
# shxrange = range(bk.nsheets)
# try:
#     sh = bk.sheet_by_name("Sheet1")
# except:
#     print "no sheet in %s named Sheet1" % fname
#  
# nrows = sh.nrows
# ncols = sh.ncols
# print "nrows %d, ncols %d" % (nrows,ncols)
#   
#  
# w.save('result.xlsx')

# import scipy.ndimage as ndimage
# import matplotlib.pyplot as plt
# import scipy.ndimage.filters as filters
# fig = plt.figure()
# plt.gray()  # show the filtered result in grayscale
# ax1 = fig.add_subplot(221)  # left side
# ax2 = fig.add_subplot(222)  # right side
# ax3 = fig.add_subplot(223)  # left side
# ax4 = fig.add_subplot(224)  # right side
# wb = Workbook()
# dest_filename = 'result1.xlsx'
# ws1 = wb.active
# ws1.title = "Sheet_1"
# wb.save(filename = dest_filename)
# fname = "result1.xlsx"
# wb = load_workbook(filename = fname)
# sheets = wb.get_sheet_names()
# count = len(sheets)
# index = count+1
# new_ws = wb.create_sheet('Sheet_'+str(index))
# wb.save(fname)


# fname = "result1.xlsx"  
# oldb = xlrd.open_workbook(fname)
# newb = xlwt.Workbook()
# count = len(oldb.sheets()) 
# print(count)
# allSheets = []
# for i in range(count):
#     allSheets.append(oldb.sheet_by_index(i))
# newb._Workbook__worksheets = allSheets
# index = count+1
# ws = newb.add_sheet('Sheet_'+str(index))
# ws.write(0,0, 'Recall')
# newb.save(fname)
# #
# def gaussian_kernel(h, w, sigma_h, sigma_w):
#     yx = np.mgrid[-h//2:h//2,-w//2:w//2]**2
#     return np.exp(-yx[0,:,:] / sigma_h**2 - yx[1,:,:] / sigma_w**2).astype('float32')
#
#
# cv2.imshow('result1',gaussian_kernel(150, 150, 25, 25))
# cv2.imshow('result2',gaussian_kernel(80, 80, 10, 10))
# cv2.waitKey()

# json_file = 'annotations/all_patient.json'
# with open(json_file) as r:
#     file_p = json.load(r)
#     for i in range(0,len(file_p)):
#         for i in range(0,len(file_p[i]["joints"])):
#             x = file_p[i]["joints"][i][0]
#             file_p[i]["joints"][i][0] = file_p[i]["joints"][i][1]
#             file_p[i]["joints"][i][1] = x
#
# with open(json_file, 'w') as w:
#     json.dump(file_p, w)

def gaussian_kernel(h, w, sigma_h, sigma_w):
    yx = np.mgrid[-h // 2:h // 2, -w // 2:w // 2] ** 2
    return np.exp(-yx[0, :, :] / sigma_h ** 2 - yx[1, :, :] / sigma_w ** 2)

def gen_kernel(score_map,border=400, sigma_h = 10, sigma_w = 10):
    kernal = gaussian_kernel(score_map.shape[0]+border, score_map.shape[1]+border, sigma_h, sigma_w)
    y, x = np.unravel_index(np.argmax(score_map), [len(score_map), len(score_map[0])])
    dh, dw = score_map.shape[0] // 2, score_map.shape[1] // 2
    y0, x0, y1, x1 = np.array([dh - y, dw - x, 3*dh - y, 3*dw - x]) + border // 2
    return kernal[y0:y1, x0:x1]

def prepare_centered_annotation(annotations, centers, H, W, train_depth, num_parts=14):
    print(annotations.shape)
    print(centers.shape)
    single_annotation_maps = np.ndarray((1, H, W, 14), dtype=np.float32)
    single_centered_resized_maps = np.ndarray((46, 38, 15), dtype=np.float32)
    annotation_maps = np.ndarray((len(annotations), 46, 38, 15), dtype=np.float32)
    for mid in range(len(annotations)):
        for pid in range(num_parts):
            score_map = np.zeros((H, W))
            score_map[annotations[mid][pid][0]][annotations[mid][pid][1]] = 1
            single_annotation_maps[0,:,:,pid] = gen_kernel(score_map)
        cv2.imwrite('help/0.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0,:,:,0]*0.3))
        cv2.imwrite('help/1.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0, :, :, 1] * 0.3))
        cv2.imwrite('help/2.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0, :, :, 2] * 0.3))
        cv2.imwrite('help/3.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0, :, :, 3] * 0.3))
        cv2.imwrite('help/4.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0, :, :, 4] * 0.3))
        cv2.imwrite('help/5.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0, :, :, 5] * 0.3))
        cv2.imwrite('help/6.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0, :, :, 6] * 0.3))
        cv2.imwrite('help/7.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0, :, :, 7] * 0.3))
        cv2.imwrite('help/8.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0, :, :, 8] * 0.3))
        cv2.imwrite('help/9.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0, :, :, 9] * 0.3))
        cv2.imwrite('help/10.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0, :, :, 10] * 0.3))
        cv2.imwrite('help/11.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0, :, :, 11] * 0.3))
        cv2.imwrite('help/12.png', 255*(train_depth[mid, :, :, 0] * 0.7 + single_annotation_maps[0, :, :, 12] * 0.3))
        cv2.imwrite('help/13.png', 255*(train_depth[mid, :, :, 0] * 0.7+ single_annotation_maps[0, :, :, 13] * 0.3))

        break
    return annotation_maps


train_depth_path = './dataset/test_autoencoder_depth.npy'
train_ir_path = './dataset/test_autoencoder_ir.npy'
train_center_path = './dataset/test_autoencoder_center.npy'
train_annotation_path = './dataset/test_autoencoder_annotation.npy'
train_images = np.load(train_depth_path)
train_centers = np.load(train_center_path)
train_annotation = np.load(train_annotation_path)
train_images = train_images.astype('float32')/255.0
prepare_centered_annotation(train_annotation, train_centers, 376, 312,train_images)