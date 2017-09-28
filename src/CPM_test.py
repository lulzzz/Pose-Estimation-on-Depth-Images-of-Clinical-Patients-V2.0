from __future__ import print_function
from __future__ import division

import numpy as np
from dataset import Dataset_2, Dataset_3
import skimage.io
import skimage.transform
import scipy.ndimage as ndimage
import scipy.ndimage.filters as filters
import cv2
import tensorflow as tf
import os
from openpyxl import load_workbook

os.environ['CUDA_VISIBLE_DEVICES'] = "3"
import cpm
import json
    
def detect_objects_heatmap(heatmap):
    data = 256 * heatmap 
    data_max = filters.maximum_filter(data, 3, mode='reflect')
    maxima = (data == data_max)
    data_min = filters.minimum_filter(data, 3, mode='reflect')
    diff = ((data_max - data_min) > 0.3)
    maxima[diff == 0] = 0
    labeled, num_objects = ndimage.label(maxima)
    slices = ndimage.find_objects(labeled)
    objects = np.zeros((num_objects, 2), dtype=np.int16)
    for oid,(dy,dx) in enumerate(slices):
        objects[oid,:] = [(dy.start + dy.stop - 1)/2, (dx.start + dx.stop - 1)/2]
    return objects

def gaussian_kernel(h, w, sigma_h, sigma_w):
    yx = np.mgrid[-h//2:h//2,-w//2:w//2]**2
    return np.exp(-yx[0,:,:] / sigma_h**2 - yx[1,:,:] / sigma_w**2)


def detect_parts_heatmaps(heatmaps, centers, size, num_parts=14):
    parts = np.zeros((len(heatmaps), num_parts, 2), dtype=np.int16)
    for i in range(len(heatmaps)):
        part_hmap = skimage.transform.resize(np.clip(heatmaps[i], -1, 1), size, mode='reflect')
        for pid in range(num_parts):
            y, x = np.unravel_index(np.argmax(part_hmap[:,:,pid]), size)
            parts[i,pid] = y+centers[i][0]-size[0]//2,x+centers[i][1]-size[1]//2
    return parts

LIMBS = np.array([1, 2, 3, 4, 4, 5, 6, 7, 7, 8,]).reshape((-1,2))-1 # 9, 10, 10, 11, 12, 13, 13, 14
COLORS = [[0, 0, 255], [0, 170, 255], [0, 255, 170], [0, 255, 0], [170, 255, 0],
          [255, 170, 0], [255, 0, 0], [255, 0, 170], [170, 0, 255]]

def draw_limbs(image, parts):
    for lid, (p0, p1) in enumerate(LIMBS):
        y0, x0 = parts[p0]
        y1, x1 = parts[p1]
        cv2.circle(image, (x0,y0), 5, (0, 0, 0), -1)
        cv2.circle(image, (x1,y1), 5, (0, 0, 0), -1)
        cv2.line(image, (x0,y0), (x1,y1), COLORS[lid], 3)

def map_to_coordinates(annotations, size, num_parts=14):
    parts_coordinates = np.zeros((len(annotations), 14, 2), dtype=np.float16)
    for mid in range(len(annotations)):
        image_resized = skimage.transform.resize(annotations[mid], size, mode='reflect')
        for pid in range(num_parts):
            y, x = np.unravel_index(np.argmax(image_resized[:,:,pid]), [image_resized.shape[0], image_resized.shape[1]])
            parts_coordinates[mid][pid][0] = y
            parts_coordinates[mid][pid][1] = x
    return parts_coordinates

def calculate_distance(point_a, point_b):
    return ((point_a[0] - point_b[0]) ** 2 + (point_a[1] - point_b[1]) ** 2) ** 0.5

def evaluation(parts_truth_coordinates,parts, threshold, num_parts=8):
    parts_predicted_coordinates = parts.astype('float16')
    matrix = np.zeros((num_parts, 2), dtype=np.uint16)
    wrong = []
    for mid in range(parts_truth_coordinates.shape[0]):
        standard_distance = threshold * calculate_distance(parts_truth_coordinates[mid][0],parts_truth_coordinates[mid][1])
        for pid in range(num_parts):
            distance = calculate_distance(parts_truth_coordinates[mid][pid],parts_predicted_coordinates[mid][pid])
            if distance <= standard_distance:
                matrix[pid][0] += 1
                matrix[pid][1] += 1
            else:
                wrong.append(mid)
                matrix[pid][1] += 1
    return matrix


def write_xls(record):
    fname = "result_pose.xlsx"
    wb = load_workbook(fname)
    sheets = wb.get_sheet_names()
    count = len(sheets)
    index = count + 1
    new_ws = wb.create_sheet('Sheet_' + str(index))

    new_ws.cell(row=1, column=1).value = 'head_top'
    new_ws.cell(row=1, column=2).value = 'jaw'
    new_ws.cell(row=1, column=3).value = 'lShoulder'
    new_ws.cell(row=1, column=4).value = 'lElbow'
    new_ws.cell(row=1, column=5).value = 'lWrist'
    new_ws.cell(row=1, column=6).value = 'rShoulder'
    new_ws.cell(row=1, column=7).value = 'rElbow'
    new_ws.cell(row=1, column=8).value = 'rWrist'
    for i in range(len(record)):
        a = record[i][0]
        b = record[i][1]
        c = a/b
        new_ws.cell(row=2, column=i+1).value = a
        new_ws.cell(row=3, column=i+ 1).value = b
        new_ws.cell(row=4, column=i+1).value = c
    wb.save(fname)


def prepare_centered_depth(images, centers, border=400):
    padded_image = np.zeros((images.shape[1] + border, images.shape[2] + border, images.shape[3]))
    results = np.zeros((images.shape[0],images.shape[1], images.shape[2], images.shape[3]))
    kernels = np.zeros((images.shape[0], images.shape[1], images.shape[2], 1))
    dh, dw = images.shape[1] // 2, images.shape[2] // 2
    for i in range(len(images)):
        padded_image[border//2:-border//2,border//2:-border//2,:] = images[i]
        yc, xc = centers[i]
        y0, x0, y1, x1 = np.array([yc - dh, xc - dw, yc + dh, xc + dw]) + border // 2
        results[i, :, :, :] = padded_image[y0:y1, x0:x1,:]
        kernels[i, :, :, 0] = gaussian_kernel(images.shape[1], images.shape[2], 25, 25)
    return results, kernels



learning_rate = 0.01
training_epochs = 20
batch_size = 2
display_step = 1
examples_to_show = 10


def test(test_depth,test_centers,saved_cpm_path,saved_result):
    with tf.device('/gpu:3'):
        pose_net_path_E3 = saved_cpm_path
        tf.reset_default_graph()
        centered_images, kernels = prepare_centered_depth(test_depth, test_centers)

        with tf.variable_scope('CPM'):
            # input dims for the pose network
            N, H, W = 3, 376, 312
            pose_image_in = tf.placeholder(tf.float32, [None,H,W,3])
            pose_centermap_in = tf.placeholder(tf.float32, [None,H,W,1])
            heatmap_pose, heatmap_stage5, heatmap_stage4, heatmap_stage3, heatmap_stage2, heatmap_stage1 = cpm.inference_pose(pose_image_in, pose_centermap_in)
            print(heatmap_pose.shape)

        tf_config = tf.ConfigProto()
        tf_config.gpu_options.allow_growth = True
        tf_config.allow_soft_placement = True

        saver = tf.train.Saver()
        b_image = centered_images - 0.5
        result = []
        with tf.Session(config=tf_config) as sess:
            saver.restore(sess, pose_net_path_E3)
            total_batch = int(len(test_depth) / batch_size)
            # Loop over all batches
            for i in range(total_batch):
                if i < total_batch- 1:
                    batch_xs = b_image[i * batch_size:(i + 1) * batch_size]
                    kernel = kernels[i * batch_size:(i + 1) * batch_size]
                else:
                    batch_xs = b_image[i * batch_size:]
                    kernel = kernels[i * batch_size:]
                # Run optimization op (backprop) and cost op (to get loss value)
                hmap_person = sess.run(heatmap_pose, feed_dict={pose_image_in: batch_xs, pose_centermap_in: kernel})
                for person_maps in hmap_person:
                    result.append(person_maps)
        print('done detecting')
        parts = detect_parts_heatmaps(result, test_centers, [H, W])
        np.save(saved_result, parts)


test_depth_path = './dataset/test_autoencoder_depth.npy'
test_ir_path = './dataset/test_autoencoder_ir.npy'
test_center_path = './dataset/test_autoencoder_center.npy'
test_annotation_path = './dataset/test_autoencoder_annotation.npy'
test_merged_path = './dataset/merged_test.npy'

test_encoderdecoder_path = './result/autoencoder3.npy'



if __name__ == '__main__':
    test_images = np.load(test_encoderdecoder_path)
    test_centers = np.load(test_center_path)
    test_annotation = np.load(test_annotation_path)
    test_images = test_images.astype('float32')/255
    #kepp depth--------------------------------------
    # test_depth = np.load(test_depth_path)
    # test_depth = test_depth.astype('float32') / 255
    #--------------------------------------------
    if test_images.shape[3] == 1:
        test_images_3 = np.zeros((test_images.shape[0], test_images.shape[1], test_images.shape[2],3))
        for i in range(test_images.shape[0]):
            test_images_3[i, :, :, 0] = test_images[i,:,:,0]
            test_images_3[i, :, :, 1] = test_images[i, :, :, 0]
            test_images_3[i, :, :, 2] = test_images[i, :, :, 0]
        print(test_images_3.shape)
        test_images = test_images_3


    saved_cpm_path = 'CPM_model/pose_net.ckpt'
    saved_result = './result/E3_parts_encoder1.npy'
    test(test_images, test_centers,saved_cpm_path,saved_result)
    result= np.load(saved_result)
    i = 0

    for i in range(len(test_images)):
        if i % 2 == 0:
            draw_limbs(test_images[i], result[i])
            cv2.imwrite('result/demo/'+str(i)+'.png',test_images[i]*255)
            i += 1

    matrix_1 = evaluation(test_annotation, result, 0.1, num_parts=8)
    matrix_2 = evaluation(test_annotation, result, 0.2, num_parts=8)
    matrix_3 = evaluation(test_annotation, result, 0.3, num_parts=8)
    matrix_4 = evaluation(test_annotation, result, 0.4, num_parts=8)
    matrix_5 = evaluation(test_annotation, result, 0.5, num_parts=8)
    print('matrix_1 -------------------------------------------------------------')
    print(matrix_1)
    print('matrix_2 -------------------------------------------------------------')
    print(matrix_2)
    print('matrix_3 -------------------------------------------------------------')
    print(matrix_3)
    print('matrix_4 -------------------------------------------------------------')
    print(matrix_4)
    print('matrix_5 -------------------------------------------------------------')
    print(matrix_5)








