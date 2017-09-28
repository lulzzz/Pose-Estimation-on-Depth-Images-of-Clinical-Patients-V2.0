from __future__ import print_function

import cv2
import numpy as np
from model import get_unet, get_vgg16
from keras.callbacks import ModelCheckpoint, LearningRateScheduler
from keras import backend as K
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from xlutils.copy import copy

vector_index = {0:'left',
                1:'slightly_left',
                2:'middle',
                3:'slightly_right',
                4:'right',
                5:'sit',
                6:'empty'}

# record = {0:{0:0,1:0,2:0,3:0,4:0,5:0,6:0},
#           1:{0:0,1:0,2:0,3:0,4:0,5:0,6:0},
#           2:{0:0,1:0,2:0,3:0,4:0,5:0,6:0},
#           3:{0:0,1:0,2:0,3:0,4:0,5:0,6:0},
#           4:{0:0,1:0,2:0,3:0,4:0,5:0,6:0},
#           5:{0:0,1:0,2:0,3:0,4:0,5:0,6:0},
#           6:{0:0,1:0,2:0,3:0,4:0,5:0,6:0}}



def load_train_data():
        imgs_train = np.load('./dataset/train.npy')
        train_label = np.load('./dataset/train_labels.npy')
        return imgs_train, train_label

def load_test_data():
        imgs_test = np.load('./dataset/test.npy')
        test_label = np.load('./dataset/test_labels.npy')
        return imgs_test, test_label
    
def visualize_result(train_label_predict, test_label):
    imgs_test, test_label = load_test_data()
    np.set_printoptions(precision=2)
    for index in range(len(test_label)):
        print(train_label_predict[index])
        print(test_label[index])
        img = imgs_test[index][0]
        img  = cv2.resize(img, (190, 380),interpolation=cv2.INTER_LINEAR)
        img = cv2.convertScaleAbs(img, alpha=(300.0/2500))
        cv2.imshow('result',img)
        if cv2.waitKey() == ord('q'):
            print('next: ') 

def statistic(train_label_predict, test_label):
    record = np.zeros((7,7))
    imgs_test, test_label = load_test_data()
    for index in range(len(test_label)):   
        ground_truth =  np.argmax(test_label[index]) 
        prediction =  np.argmax(train_label_predict[index])
        record[ground_truth][prediction] += 1
    return record
    
    
def write_xls(record):  
    fname = "result1.xlsx"  
    wb = load_workbook(filename = fname)
    sheets = wb.get_sheet_names()  
    count = len(sheets) 
    index = count+1
    new_ws = wb.create_sheet('Sheet_'+str(index))

    for i in range(len(vector_index)):
        new_ws.cell(row = 1, column = i+2).value = vector_index[i]
        new_ws.cell(row = i+2, column = 1).value = vector_index[i]
    
    for i in range(len(record)):
        for j in range(len(record)):
            new_ws.cell(row = i+2, column = j+2).value = record[i][j]
            
    sum_x = np.sum(record,axis = 1)
    sum_y = np.sum(record,axis = 0)
    new_ws.cell(row = 1, column = len(record)+3).value = 'Recall'
    new_ws.cell(row = len(record)+3, column = 1).value = 'Precision'
 
    for i in range(len(record)):    
        new_ws.cell(row = i+2, column = len(record)+2).value = sum_x[i]
        recall = record[i][i]/sum_x[i]
        new_ws.cell(row = i+2, column = len(record)+3).value = recall
    for j in range(len(record)):
        new_ws.cell(row = len(record)+2, column = j+2).value = sum_y[j]
        precision = record[j][j]/sum_y[j]
        new_ws.cell(row = len(record)+3, column = j+2).value = precision
    wb.save(fname)  
    
def train():
    model = get_vgg16()
    print('Loading and preprocessing train data...')
    print('-'*30)
    imgs_train, train_label = load_train_data()
    imgs_train = imgs_train.astype('float32')
    train_label = train_label.astype('float32')
    mean = np.mean(imgs_train)  # mean for data centering
    std = np.std(imgs_train)  # std for data normalization
    imgs_train -= mean
    imgs_train /= std
#     imgs_train /= 2500.0
#     imgs_train *= 255.0
    print(imgs_train[1][0][50])
    imgs_test, test_label = load_test_data()
    imgs_test = imgs_test.astype('float32')
    test_label = test_label.astype('float32')
    mean = np.mean(imgs_test)  # mean for data centering
    std = np.std(imgs_test)  # std for data normalization

    imgs_test -= mean
    imgs_test /= std
#     imgs_test /= 2500.0
#     imgs_test *= 255.0
    print('Creating and compiling model...')
    print('-'*30)
    checkpointer = ModelCheckpoint(filepath="weights_posture.hdf5", verbose=1)#, save_best_only=True
    model.fit(imgs_train, train_label, batch_size=8, epochs=20, verbose=2, shuffle=True, 
              validation_data=(imgs_test, test_label),callbacks=[checkpointer])
    
#     imgs_test = load_test_data()
#     imgs_test = imgs_test.astype('float32')
#     imgs_test /= 2500
#     
#     train_label_predict = model.predict(imgs_test, verbose=1) 
#     visualize_result(train_label_predict)
    
def predict():
    model = get_vgg16()
    print('Loading and preprocessing train data...')
    print('-'*30)
    imgs_test, test_label = load_test_data()
    imgs_test = imgs_test.astype('float32')
    test_label = test_label.astype('float32')
    mean = np.mean(imgs_test)  # mean for data centering
    std = np.std(imgs_test)  # std for data normalization

    imgs_test -= mean
    imgs_test /= std
    print('Loading saved weights...')
    print('-'*30)
    model.load_weights('weights_posture.hdf5')
#     model.load_weights('best.hdf5')

    print('Predicting masks on play data...')
    print('-'*30)
    train_label_predict = model.predict(imgs_test, verbose=1)
    
#     visualize_result(train_label_predict, test_label)
    record = statistic(train_label_predict, test_label)
    write_xls(record)
    
if __name__ == '__main__':
#     train()
    predict()